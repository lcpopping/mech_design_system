"""Standard parts calculation routes."""
from flask import Blueprint, request, jsonify, send_file
from backend.services.bearing_service import BearingService
from backend.services.servo_motor_service import ServoMotorService
from backend.services.key_service import KeyService
from io import BytesIO
from datetime import datetime

standard_parts_bp = Blueprint('standard_parts', __name__)


try:
    from openpyxl import Workbook
    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False


# ==================== Bearing Routes ====================

@standard_parts_bp.route('/bearings/types', methods=['GET'])
def get_bearing_types():
    types = BearingService.get_bearing_types()
    return jsonify({'success': True, 'data': types})


@standard_parts_bp.route('/bearings', methods=['GET'])
def get_bearings():
    bearing_type = request.args.get('type')
    bore_min = request.args.get('bore_min', type=float)
    bore_max = request.args.get('bore_max', type=float)
    load_min = request.args.get('load_min', type=float)
    speed_max = request.args.get('speed_max', type=float)

    bearings = BearingService.query_bearings(
        bearing_type=bearing_type,
        bore_min=bore_min,
        bore_max=bore_max,
        load_min=load_min,
        speed_max=speed_max
    )
    return jsonify({'success': True, 'data': bearings})


@standard_parts_bp.route('/bearings/<int:bearing_id>', methods=['GET'])
def get_bearing(bearing_id):
    bearing = BearingService.get_bearing_by_id(bearing_id)
    if not bearing:
        return jsonify({'success': False, 'error': 'Bearing not found'}), 404
    return jsonify({'success': True, 'data': bearing})


# ==================== Bearing Selection Wizard ====================

@standard_parts_bp.route('/bearings/wizard/types', methods=['GET'])
def wizard_get_types():
    types = BearingService.get_bearing_types()
    return jsonify({
        'success': True,
        'data': {
            'types': types,
            'step_info': {'current_step': 1, 'total_steps': 4, 'title': '选择轴承类型'}
        }
    })


@standard_parts_bp.route('/bearings/wizard/query', methods=['POST'])
def wizard_query_candidates():
    data = request.get_json()

    bearing_type = data.get('bearing_type', 'deep_groove')
    fr = data.get('fr', 0)
    fa = data.get('fa', 0)
    speed = data.get('speed', 0)
    temperature = data.get('temperature', 70)
    lubrication = data.get('lubrication', 'grease')
    required_life = data.get('required_life_hours', 10000)
    reliability = data.get('reliability', 0.9)
    bore_max = data.get('bore_max')
    contamination = data.get('contamination_factor', 1.0)

    bearing_id = data.get('bearing_id')
    if bearing_id:
        result = BearingService.calculate_full_spec(
            bearing_id=bearing_id,
            fr=fr, fa=fa, speed=speed,
            temperature=temperature, lubrication=lubrication,
            reliability=reliability
        )
        if 'error' in result:
            return jsonify({'success': False, 'error': result['error']}), 400
        return jsonify({
            'success': True,
            'data': {
                'specifications': result,
                'step_info': {'current_step': 4, 'total_steps': 4, 'title': '选型结果'}
            }
        })

    result = BearingService.intelligent_select(
        bearing_type=bearing_type,
        fr=fr, fa=fa, speed=speed,
        temperature=temperature, lubrication=lubrication,
        required_life_hours=required_life,
        reliability=reliability,
        bore_max=bore_max,
        contamination_factor=contamination
    )

    if 'error' in result:
        return jsonify({'success': False, 'error': result['error']}), 400

    return jsonify({
        'success': True,
        'data': {
            'candidates': result.get('candidates', [])[:5],
            'recommended': result.get('recommended'),
            'required_capacity': result.get('required_capacity_N'),
            'equivalent_load': result.get('equivalent_load_P'),
            'step_info': {'current_step': 3, 'total_steps': 4, 'title': '选择轴承型号'}
        }
    })


@standard_parts_bp.route('/bearings/wizard/select', methods=['POST'])
def wizard_select():
    data = request.get_json()

    bearing_id = data.get('bearing_id')
    fr = data.get('fr', 0)
    fa = data.get('fa', 0)
    speed = data.get('speed', 0)
    temperature = data.get('temperature', 70)
    lubrication = data.get('lubrication', 'grease')
    reliability = data.get('reliability', 0.9)

    result = BearingService.calculate_full_spec(
        bearing_id=bearing_id,
        fr=fr, fa=fa, speed=speed,
        temperature=temperature, lubrication=lubrication,
        reliability=reliability
    )

    if 'error' in result:
        return jsonify({'success': False, 'error': result['error']}), 400

    return jsonify({
        'success': True,
        'data': {
            'specifications': result,
            'step_info': {'current_step': 4, 'total_steps': 4, 'title': '选型结果'}
        }
    })


# ==================== Export Excel ====================

@standard_parts_bp.route('/bearings/export/excel', methods=['POST'])
def export_bearing_selection_excel():
    """Export bearing selection result to Excel file."""
    if not HAS_EXCEL:
        return jsonify({'success': False, 'error': 'openpyxl not installed'}), 500

    data = request.get_json()
    spec = data.get('specifications', {})
    params = data.get('params', {})

    wb = Workbook()
    ws = wb.active
    ws.title = '轴承选型报告'

    # Header
    ws['A1'] = '轴承选型计算报告'
    ws['A1'].font = ws['A1'].font.copy(bold=True, size=14)
    ws.merge_cells('A1:D1')

    ws['A2'] = f'生成时间: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    ws['A3'] = ''

    # Basic info
    b = spec.get('bearing', {})
    ws['A4'] = '【轴承型号】'
    ws['A4'].font = ws['A4'].font.copy(bold=True)
    ws['B4'] = b.get('model', '-')
    ws['A5'] = '类型'
    ws['B5'] = spec.get('bearing_type', '-')
    ws['A6'] = '内径 (mm)'
    ws['B6'] = b.get('bore_diameter', '-')
    ws['C6'] = '外径 (mm)'
    ws['D6'] = b.get('outer_diameter', '-')
    ws['A7'] = '宽度 (mm)'
    ws['B7'] = b.get('width', '-')
    ws['C7'] = '重量 (kg)'
    ws['D7'] = b.get('weight', '-')

    # Load ratings
    ws['A9'] = '【额定载荷】'
    ws['A9'].font = ws['A9'].font.copy(bold=True)
    ws['A10'] = '动载荷C (N)'
    ws['B10'] = b.get('dynamic_load_rating', '-')
    ws['A11'] = '静载荷C0 (N)'
    ws['B11'] = b.get('static_load_rating', '-')
    ws['C11'] = '极限转速 (rpm)'
    ws['D11'] = b.get('max_speed', '-')

    # Working conditions
    ml = spec.get('modified_life', {})
    ws['A13'] = '【工况参数】'
    ws['A13'].font = ws['A13'].font.copy(bold=True)
    ws['A14'] = '径向载荷Fr (N)'
    ws['B14'] = params.get('fr', 0)
    ws['C14'] = '轴向载荷Fa (N)'
    ws['D14'] = params.get('fa', 0)
    ws['A15'] = '转速 (rpm)'
    ws['B15'] = params.get('speed', 0)
    ws['C15'] = '温度 (°C)'
    ws['D15'] = params.get('temperature', 70)
    ws['A16'] = '润滑方式'
    ws['B16'] = params.get('lubrication', 'grease')
    ws['C16'] = '可靠度'
    ws['D16'] = f"{params.get('reliability', 0.9) * 100}%"

    # Results
    ws['A18'] = '【计算结果】'
    ws['A18'].font = ws['A18'].font.copy(bold=True)
    ws['A19'] = '当量动载荷P (N)'
    ws['B19'] = spec.get('equivalent_load_P', 0)
    ws['A20'] = 'L10基本寿命 (小时)'
    ws['B20'] = ml.get('l10_life_hours', 0)
    ws['C20'] = 'L10m修正寿命 (小时)'
    ws['D20'] = ml.get('l10m_life_hours', 0)
    ws['A21'] = '可靠度因子a1'
    ws['B21'] = ml.get('reliability_factor_a1', 1)
    ws['C21'] = '综合因子aISO'
    ws['D21'] = ml.get('composite_factor_a_iso', 0)

    # Static check
    sc = spec.get('static_check', {})
    ws['A23'] = '【静载荷校核】'
    ws['A23'].font = ws['A23'].font.copy(bold=True)
    ws['A24'] = '当量静载荷P0 (N)'
    ws['B24'] = sc.get('equivalent_static_load', 0)
    ws['C24'] = '安全系数s0'
    ws['D24'] = sc.get('safety_factor', 0)
    ws['A25'] = '校核结果'
    ws['B25'] = '通过' if sc.get('passed') else '未通过'

    # Stiffness
    st = spec.get('stiffness', {})
    ws['A27'] = '【刚度参数】'
    ws['A27'].font = ws['A27'].font.copy(bold=True)
    ws['A28'] = '径向刚度 (N/μm)'
    ws['B28'] = st.get('radial_stiffness_N_per_um', 0)
    ws['C28'] = '轴向刚度 (N/μm)'
    ws['D28'] = st.get('axial_stiffness_N_per_um', 0)

    # Set column widths
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 18

    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"轴承选型_{b.get('model', 'report')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


# ==================== Legacy API ====================

@standard_parts_bp.route('/bearings/select', methods=['POST'])
def select_bearing():
    data = request.get_json()
    result = BearingService.intelligent_select(
        bearing_type=data.get('bearing_type', 'deep_groove'),
        fr=data.get('load', 0),
        fa=data.get('fa', 0),
        speed=data.get('speed', 0),
        temperature=data.get('temperature', 70),
        lubrication=data.get('lubrication', 'grease'),
        required_life_hours=data.get('life_required', 10000),
        reliability=data.get('reliability', 0.9)
    )
    if 'error' in result:
        return jsonify({'success': False, 'error': result['error']}), 400

    recommended = result.get('recommended')
    if not recommended:
        return jsonify({'success': False, 'error': 'No suitable bearing'}), 400

    return jsonify({
        'success': True,
        'data': {
            'bearing': recommended.get('bearing'),
            'l10_life_hours': recommended.get('l10m_life_hours', 0),
            'safety_factor': recommended.get('safety_factor', 0),
            'static_pass': recommended.get('static_check', {}).get('passed', False)
        }
    })


# ==================== Servo Motor Routes ====================

@standard_parts_bp.route('/servo-motors/types', methods=['GET'])
def get_servo_motor_types():
    types = ServoMotorService.get_servo_motor_types()
    return jsonify({'success': True, 'data': types})


@standard_parts_bp.route('/servo-motors', methods=['GET'])
def get_servo_motors():
    power_min = request.args.get('power_min', type=float)
    power_max = request.args.get('power_max', type=float)
    torque_min = request.args.get('torque_min', type=float)
    torque_max = request.args.get('torque_max', type=float)
    motors = ServoMotorService.query_servo_motors(
        power_min_w=power_min, power_max_w=power_max,
        torque_min_nm=torque_min, torque_max_nm=torque_max
    )
    return jsonify({'success': True, 'data': motors})


@standard_parts_bp.route('/servo-motors/<int:motor_id>', methods=['GET'])
def get_servo_motor(motor_id):
    motor = ServoMotorService.get_servo_motor_by_id(motor_id)
    if not motor:
        return jsonify({'success': False, 'error': 'Servo motor not found'}), 404
    return jsonify({'success': True, 'data': motor})


@standard_parts_bp.route('/servo-motors/select', methods=['POST'])
def select_servo_motor():
    data = request.get_json()
    load_mass_kg = data.get('load_mass_kg', 0)
    speed_ms = data.get('speed_ms', 0)
    speed_rpm = data.get('speed_rpm', 0)
    motion_type = data.get('motion_type', 'linear')
    guide_type = data.get('guide_type', 'linear_guide')
    wheel_radius_m = data.get('wheel_radius_m', 0.05)
    screw_lead_m = data.get('screw_lead_m', 0.005)
    safety_factor = data.get('safety_factor', 1.2)
    load_type = data.get('load_type', 'h')
    transmission_type = data.get('transmission_type', 'direct')
    reduction_ratio = data.get('reduction_ratio', 1.0)
    friction_coefficient = data.get('friction_coefficient', 0.1)
    duty_cycle = data.get('duty_cycle', 'S1')
    inertia_matching_limit = data.get('inertia_matching_limit', 5)

    result = ServoMotorService.intelligent_select(
        load_mass_kg=load_mass_kg,
        speed_ms=speed_ms,
        speed_rpm=speed_rpm,
        motion_type=motion_type,
        guide_type=guide_type,
        wheel_radius_m=wheel_radius_m,
        screw_lead_m=screw_lead_m,
        safety_factor=safety_factor,
        load_type=load_type,
        transmission_type=transmission_type,
        reduction_ratio=reduction_ratio,
        friction_coefficient=friction_coefficient,
        duty_cycle=duty_cycle,
        inertia_matching_limit=inertia_matching_limit
    )

    if 'error' in result:
        return jsonify({'success': False, 'error': result['error']}), 400

    return jsonify({
        'success': True,
        'data': result
    })


# ==================== Key Routes ====================

@standard_parts_bp.route('/keys/check', methods=['POST'])
def check_key_strength():
    """Check key strength under given torque."""
    data = request.get_json()
    torque_Nm = data.get('torque_Nm', 0)
    shaft_diameter_mm = data.get('shaft_diameter_mm', 0)
    key_width_mm = data.get('key_width_mm', 0)
    key_height_mm = data.get('key_height_mm', 0)
    key_length_mm = data.get('key_length_mm', 0)
    material = data.get('material', 'carbon_steel')
    safety_factor = data.get('safety_factor', 1.0)

    result = KeyService.check_key_strength(
        torque_Nm=torque_Nm,
        shaft_diameter_mm=shaft_diameter_mm,
        key_width_mm=key_width_mm,
        key_height_mm=key_height_mm,
        key_length_mm=key_length_mm,
        material=material,
        safety_factor=safety_factor
    )

    return jsonify({'success': True, 'data': result})


@standard_parts_bp.route('/keys/recommend-length', methods=['POST'])
def recommend_key_length():
    """Recommend minimum key length based on constraint."""
    data = request.get_json()
    torque_Nm = data.get('torque_Nm', 0)
    shaft_diameter_mm = data.get('shaft_diameter_mm', 0)
    key_width_mm = data.get('key_width_mm', 0)
    key_height_mm = data.get('key_height_mm', 0)
    material = data.get('material', 'carbon_steel')
    safety_factor = data.get('safety_factor', 1.0)
    constraint = data.get('constraint', 'shear')

    result = KeyService.recommend_key_length(
        torque_Nm=torque_Nm,
        shaft_diameter_mm=shaft_diameter_mm,
        key_width_mm=key_width_mm,
        key_height_mm=key_height_mm,
        material=material,
        safety_factor=safety_factor,
        constraint=constraint
    )

    return jsonify({'success': True, 'data': result})


@standard_parts_bp.route('/keys/materials', methods=['GET'])
def get_key_materials():
    """Get available key materials."""
    materials = KeyService.get_material_options()
    return jsonify({'success': True, 'data': materials})


@standard_parts_bp.route('/keys/standard-sizes', methods=['GET'])
def get_standard_key_sizes():
    """Get standard key sizes for given shaft diameter."""
    shaft_diameter_mm = request.args.get('shaft_diameter_mm', type=float)
    if not shaft_diameter_mm:
        return jsonify({'success': False, 'error': 'shaft_diameter_mm is required'}), 400

    sizes = KeyService.get_standard_key_sizes(shaft_diameter_mm)
    return jsonify({'success': True, 'data': sizes})


